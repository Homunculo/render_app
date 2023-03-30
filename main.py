from flask import Flask, request
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        wb = load_workbook(filename=file)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            print(row)
        return "File uploaded and printed!"
    return '''
    <html>
        <body>
            <h1>Upload Excel File</h1>
            <form method="post" enctype="multipart/form-data">
                <input type="file" name="file">
                <input type="submit" value="Upload">
            </form>
        </body>
    </html>
    '''

if __name__ == '__main__':
    app.run(debug=True, port=80)
